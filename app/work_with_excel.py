import os
import csv
import zipfile
from io import StringIO, BytesIO
from datetime import datetime
from datetime import date as d_date
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import load_workbook
from openpyxl import styles
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse


def ips_import(file) -> list:
    result = []
    wb_obj = load_workbook(file, data_only=True)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    for row in range(2, m_row + 1):
        result.append(sheet_obj.cell(row=row, column=1).value)
    return result


def export_search_result(node_list, ports, modules, domains, vulns, relations):
    workbook_list = []
    if node_list:
        workbook = Workbook()
        worksheet = workbook.active
        tdate = d_date.today().strftime("%d.%m.%Y")
        worksheet.title = '' + tdate
        header = ['ID', 'Домен/поддомен', 'IP', 'Родительский домен', 'Найдено модулями', 'Открытые порты', 'Баннер',
                  'Уязвимости', 'Есть в БД Периметра', 'Новый сервис', 'Дата добавления']
        new_nodes = []

        # Обработка Total
        # Перевод вложенных tuple в списки
        for i in node_list:
            new_nodes += [list(i)]

        # Формирование шапки файла
        for col_num, h in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = Font(name='Times New Roman', size='12', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.value = h

        # Заполнение таблицы
        for row_num, line in enumerate(new_nodes, 2):
            for col_num, cell_value in enumerate(line, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                if col_num in [3, 5, 7, 8]:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Times New Roman', size='12', bold=False)
                cell.value = cell_value
        worksheet.column_dimensions['A'].width = 5
        worksheet.column_dimensions['B'].width = 30
        worksheet.column_dimensions['C'].width = 17
        worksheet.column_dimensions['D'].width = 23
        worksheet.column_dimensions['E'].width = 23
        worksheet.column_dimensions['F'].width = 21
        worksheet.column_dimensions['G'].width = 60
        worksheet.column_dimensions['H'].width = 50
        worksheet.column_dimensions['I'].width = 23
        worksheet.column_dimensions['J'].width = 21
        worksheet.column_dimensions['K'].width = 21

        # Добавление фильтра
        worksheet.auto_filter.ref = 'A1:K' + str(len(new_nodes) + 1)

        # Обработка Ports
        worksheet = workbook.create_sheet(title='Ports')
        worksheet = workbook.get_sheet_by_name('Ports')
        header = ['№', 'Имя']
        # Формирование шапки файла
        for col_num, h in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = Font(name='Times New Roman', size='12', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.value = h

        # Заполнение таблицы
        for row_num, line in enumerate(ports, 2):
            for col_num, cell_value in enumerate(line, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Times New Roman', size='12', bold=False)
                cell.value = cell_value
        worksheet.column_dimensions['A'].width = 7
        worksheet.column_dimensions['B'].width = 15
        worksheet.auto_filter.ref = 'A1:B' + str(len(ports) + 1)

        # Обработка Relations
        worksheet = workbook.create_sheet(title='Relations')
        worksheet = workbook.get_sheet_by_name('Relations')
        header = ['ID', 'Домен/поддомен', 'Запрос', 'Модуль']
        # Формирование шапки файла
        for col_num, h in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = Font(name='Times New Roman', size='12', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.value = h

        # Заполнение таблицы
        for row_num, line in enumerate(relations, 2):
            for col_num, cell_value in enumerate(line, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Times New Roman', size='12', bold=False)
                cell.value = cell_value
        worksheet.column_dimensions['A'].width = 7
        worksheet.column_dimensions['B'].width = 25
        worksheet.column_dimensions['C'].width = 25
        worksheet.column_dimensions['D'].width = 25
        worksheet.auto_filter.ref = 'A1:D' + str(len(relations) + 1)

        # Обработка Vulns
        worksheet = workbook.create_sheet(title='Vulns')
        worksheet = workbook.get_sheet_by_name('Vulns')
        header = ['CVE', 'Оценка CVSS', 'Описание']
        # Формирование шапки файла
        for col_num, h in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = Font(name='Times New Roman', size='12', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.value = h

        # Заполнение таблицы
        for row_num, line in enumerate(vulns, 2):
            for col_num, cell_value in enumerate(line, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.font = Font(name='Times New Roman', size='12', bold=False)
                cell.value = cell_value
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 200
        worksheet.auto_filter.ref = 'A1:C' + str(len(vulns) + 1)

        # Обработка Modules
        worksheet = workbook.create_sheet(title='Modules')
        worksheet = workbook.get_sheet_by_name('Modules')
        header = ['ID', 'Имя', 'Уровень', 'Описание']
        # Формирование шапки файла
        for col_num, h in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = Font(name='Times New Roman', size='12', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.value = h

        # Заполнение таблицы
        for row_num, line in enumerate(modules, 2):
            for col_num, cell_value in enumerate(line, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Times New Roman', size='12', bold=False)
                cell.value = cell_value
        worksheet.column_dimensions['A'].width = 5
        worksheet.column_dimensions['B'].width = 17
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 125
        worksheet.auto_filter.ref = 'A1:D' + str(len(modules) + 1)

        # Обработка Domains
        worksheet = workbook.create_sheet(title='Domains')
        worksheet = workbook.get_sheet_by_name('Domains')
        header = ['№', 'Имя']
        # Формирование шапки файла
        for col_num, h in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = Font(name='Times New Roman', size='12', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.value = h

        # Заполнение таблицы
        for row_num, line in enumerate(domains, 2):
            cell = worksheet.cell(row=row_num, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Times New Roman', size='12', bold=False)
            cell.value = row_num - 1
            cell = worksheet.cell(row=row_num, column=2)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Times New Roman', size='12', bold=False)
            cell.value = line

        worksheet.column_dimensions['A'].width = 5
        worksheet.column_dimensions['B'].width = 25
        worksheet.auto_filter.ref = 'A1:B' + str(len(domains) + 1)

        file_name = 'Результаты поиска {}.xlsx'.format(tdate)
        workbook_list.append((file_name, BytesIO(save_virtual_workbook(workbook))))

        # Генерация архива
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
            for file_name, data in workbook_list:
                zip_file.writestr(file_name, data.getvalue())

        resp = HttpResponse(zip_buffer.getvalue(), content_type="application/x-zip-compressed")
        # ..and correct content-disposition

        resp['Content-Disposition'] = 'attachment; filename=scanresult-{date}.zip'.format(
            date=datetime.now().strftime("%d.%m.%Y"),
        )
        return resp
